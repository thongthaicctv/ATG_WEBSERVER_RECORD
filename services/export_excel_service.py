# services/export_excel_service.py
# -*- coding: utf-8 -*-

from io import BytesIO
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="0B5ED7")
HEADER_FONT = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Times New Roman", size=13)
TITLE_FONT = Font(name="Times New Roman", size=16, bold=True, color="0B2545")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def safe_value(value):
    if value is None:
        return ""

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return value


def _save_to_memory(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_title(ws, title, subtitle="", max_col=8):
    end_col = get_column_letter(max_col)

    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = NORMAL_FONT
    ws["A2"].alignment = CENTER


def _format_sheet(ws, header_row=4):
    ws.freeze_panes = f"A{header_row + 1}"

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    for cell in ws[header_row]:
        if isinstance(cell, MergedCell):
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0

        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 12), 60)


def export_employee_report(rows, date_from="", date_to=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao nhan vien"

    headers = [
        "STT",
        "Ngày",
        "Mã NV",
        "Tên nhân viên",
        "Tổng đơn",
        "Tổng video",
        "Tổng thời lượng giây",
        "Tổng thời lượng phút",
    ]

    subtitle = f"Từ ngày: {date_from or '...'}  đến ngày: {date_to or '...'}"
    _write_title(ws, "BÁO CÁO NĂNG SUẤT NHÂN VIÊN", subtitle, max_col=len(headers))

    ws.append([])
    ws.append(headers)

    for idx, r in enumerate(rows, start=1):
        total_seconds = int(r.get("total_duration_seconds") or 0)
        total_minutes = round(total_seconds / 60, 2)

        ws.append([
            idx,
            safe_value(r.get("work_date")),
            safe_value(r.get("employee_code")),
            safe_value(r.get("employee_name")),
            safe_value(r.get("total_orders") or 0),
            safe_value(r.get("total_videos") or 0),
            total_seconds,
            total_minutes,
        ])

    _format_sheet(ws, header_row=4)
    return _save_to_memory(wb)


def export_video_report(rows, date_from="", date_to=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao video"

    headers = [
        "STT",
        "Thời gian",
        "Mã đơn",
        "Loại",
        "Scanner",
        "Camera",
        "Nhân viên",
        "File",
        "Dung lượng MB",
        "Thời lượng giây",
        "Kết quả",
    ]

    subtitle = f"Từ ngày: {date_from or '...'}  đến ngày: {date_to or '...'}"
    _write_title(ws, "BÁO CÁO VIDEO THEO NGÀY", subtitle, max_col=len(headers))

    ws.append([])
    ws.append(headers)

    for idx, r in enumerate(rows, start=1):
        file_size_mb = round((r.get("file_size") or 0) / 1024 / 1024, 2)

        ws.append([
            idx,
            safe_value(r.get("start_time") or r.get("created_at")),
            safe_value(r.get("order_code")),
            safe_value(r.get("session_type") or r.get("video_type")),
            safe_value(r.get("scanner_id")),
            safe_value(f"{r.get('camera_id') or ''} - {r.get('camera_name') or ''}"),
            safe_value(r.get("employee_name") or r.get("employee_code")),
            safe_value(r.get("file_name")),
            file_size_mb,
            safe_value(r.get("duration_seconds") or 0),
            safe_value(r.get("result")),
        ])

    _format_sheet(ws, header_row=4)
    return _save_to_memory(wb)


def export_missing_video_report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Video loi thieu file"

    headers = [
        "STT",
        "ID",
        "Mã đơn",
        "Camera",
        "File name",
        "File path",
        "Dung lượng DB",
        "Thời lượng",
        "Ngày tạo",
    ]

    _write_title(
        ws,
        "BÁO CÁO VIDEO LỖI / THIẾU FILE",
        f"Xuất lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        max_col=len(headers),
    )

    ws.append([])
    ws.append(headers)

    for idx, r in enumerate(rows, start=1):
        ws.append([
            idx,
            safe_value(r.get("id")),
            safe_value(r.get("order_code")),
            safe_value(f"{r.get('camera_id') or ''} - {r.get('camera_name') or ''}"),
            safe_value(r.get("file_name")),
            safe_value(r.get("file_path")),
            safe_value(r.get("file_size") or 0),
            safe_value(r.get("duration_seconds") or 0),
            safe_value(r.get("created_at")),
        ])

    _format_sheet(ws, header_row=4)
    ws.column_dimensions["F"].width = 75
    return _save_to_memory(wb)


def export_order_status_report(rows, date_from="", date_to=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trang thai don hang"

    headers = [
        "STT",
        "Mã đơn",
        "Loại đơn",
        "Khách hàng",
        "SĐT",
        "Số video",
        "Đóng hàng",
        "Check băng truyền",
        "Vận chuyển",
        "Mã vận đơn",
        "thời gian cập nhật vận chuyển",
    ]

    subtitle = f"Từ ngày: {date_from or '...'}  đến ngày: {date_to or '...'}"
    _write_title(ws, "BÁO CÁO TRẠNG THÁI ĐƠN HÀNG", subtitle, max_col=len(headers))

    ws.append([])
    ws.append(headers)

    for idx, r in enumerate(rows, start=1):
        ws.append([
            idx,
            safe_value(r.get("order_code")),
            safe_value(r.get("order_type")),
            safe_value(r.get("customer_name")),
            safe_value(r.get("customer_phone")),
            safe_value(r.get("video_count") or 0),
            safe_value(r.get("packing_status")),
            safe_value(r.get("conveyor_status")),
            safe_value(r.get("shipping_status")),
            safe_value(r.get("tracking_code")),
            safe_value(r.get("shipping_updated_at")),
        ])

    _format_sheet(ws, header_row=4)
    return _save_to_memory(wb)


def export_order_status_split_report(ecom_rows, wholesale_rows, date_from="", date_to=""):
    wb = Workbook()
    subtitle = f"Từ ngày: {date_from or '...'}  đến ngày: {date_to or '...'}"

    ws = wb.active
    ws.title = "Trang thai don hang Ecom"
    _write_order_status_sheet(
        ws=ws,
        title="BÁO CÁO TRẠNG THÁI ĐƠN HÀNG",
        subtitle=subtitle,
        rows=ecom_rows,
        is_wholesale=False,
    )

    ws2 = wb.create_sheet("Trang thai don hang Si")
    _write_order_status_sheet(
        ws=ws2,
        title="BÁO CÁO TRẠNG THÁI ĐƠN HÀNG",
        subtitle=subtitle,
        rows=wholesale_rows,
        is_wholesale=True,
    )

    return _save_to_memory(wb)


def _write_order_status_sheet(ws, title, subtitle, rows, is_wholesale=False):
    if is_wholesale:
        headers = [
            "STT",
            "Mã đơn",
            "Loại đơn",
            "Khách hàng",
            "SĐT",
            "Số video",
            "Đóng hàng",
            "Thời gian đóng",
            "Chi tiết kiện nhỏ",
            "nhân viên đóng",
            "nhân viên giao",
            "Vận chuyển",
            "Mã vận đơn",
            "thời gian cập nhật vận chuyển",
        ]
    else:
        headers = [
            "STT",
            "Mã đơn",
            "Loại đơn",
            "Khách hàng",
            "SĐT",
            "Số video",
            "Đóng hàng",
            "Thời gian đóng",
            "Nhân viên",
            "Check băng truyền",
            "Vận chuyển",
            "Mã vận đơn",
            "thời gian cập nhật vận chuyển",
        ]

    _write_title(ws, title, subtitle, max_col=len(headers))
    ws.append([])
    ws.append(headers)

    for idx, r in enumerate(rows, start=1):
        base = [
            idx,
            safe_value(r.get("order_code")),
            safe_value(r.get("order_type")),
            safe_value(r.get("customer_name")),
            safe_value(r.get("customer_phone")),
            safe_value(r.get("video_count") or 0),
            safe_value(r.get("packing_status")),
            safe_value(r.get("packed_at")),
        ]

        if is_wholesale:
            ws.append(base + [
                safe_value(r.get("small_package_detail") or "Chưa có dữ liệu kiện nhỏ"),
                safe_value(r.get("packing_employee")),
                safe_value(r.get("delivery_employee")),
                safe_value(r.get("shipping_status")),
                safe_value(r.get("tracking_code")),
                safe_value(r.get("shipping_updated_at")),
            ])
        else:
            ws.append(base + [
                safe_value(r.get("packing_employee") or r.get("employee_name") or r.get("employee_code")),
                safe_value(r.get("conveyor_status")),
                safe_value(r.get("shipping_status")),
                safe_value(r.get("tracking_code")),
                safe_value(r.get("shipping_updated_at")),
            ])

    _format_sheet(ws, header_row=4)

    widths = {
        "A": 5,
        "B": 24,
        "C": 18 if is_wholesale else 12,
        "D": 22,
        "F": 13,
        "G": 13,
        "H": 22,
        "I": 38 if is_wholesale else 26,
        "J": 16 if is_wholesale else 38,
        "K": 16 if is_wholesale else 13,
        "L": 13,
        "M": 22 if not is_wholesale else 13,
        "N": 22,
    }
    for col, width in widths.items():
        if column_index_from_string(col) <= ws.max_column:
            ws.column_dimensions[col].width = width

    left_columns = (9, 10, 11) if is_wholesale else (10,)
    for col_idx in left_columns:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = LEFT

    for row_idx in range(5, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in left_columns]
        line_count = max((str(value).count("\n") + 1 for value in values if value), default=1)
        if line_count > 1:
            ws.row_dimensions[row_idx].height = max(22, line_count * 18)
